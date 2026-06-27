import os
import sys
import platform
import subprocess
import time
import pandas as pd
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QVBoxLayout, QLabel, QPushButton,
    QFileDialog, QWidget, QMessageBox, QCheckBox,
)
from PyQt5.QtGui import QIcon
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

COLUMNS = [
    'RefDes', 'Part Number', 'Description', 'Producer Abbreviation',
    'Producer Part Name', 'Validity Flag', 'Validity Flag Remarks',
    'Case Size', 'Collective Numbers', 'Quantity', 'Placement Side Information',
]
OUTPUT_DIR = 'Result'


def open_file(path):
    if platform.system() == 'Windows':
        os.startfile(path)
    elif platform.system() == 'Darwin':
        subprocess.run(['open', path])
    else:
        subprocess.run(['xdg-open', path])


class BOMComparisonApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("BOM Comparison Tool")
        self.setWindowIcon(QIcon("bosch.ico"))
        screen = QApplication.primaryScreen().availableGeometry()
        w, h = 600, 200
        self.setGeometry((screen.width() - w) // 2, (screen.height() - h) // 2, w, h)
        self.file1 = None
        self.file2 = None
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout()

        self.label_file1 = QLabel("First BOM File: Not Selected")
        self.label_file2 = QLabel("Second BOM File: Not Selected")
        layout.addWidget(self.label_file1)
        layout.addWidget(self.label_file2)

        layout.addWidget(self._button("Select First BOM File", self.select_file1))
        layout.addWidget(self._button("Select Second BOM File", self.select_file2))

        self.cb_zero_qty = QCheckBox("Remove Zero Quantity in Both BOMs")
        self.cb_new_zero = QCheckBox("Remove New Zero Quantity in BOM2")
        self.cb_dep_zero = QCheckBox("Remove Deprecated Zero Quantity in BOM1")
        self.cb_unchanged = QCheckBox("Remove Unchanged Items")

        for cb in [self.cb_zero_qty, self.cb_new_zero, self.cb_dep_zero, self.cb_unchanged]:
            layout.addWidget(cb)

        layout.addWidget(self._button("Run Comparison", self.run_comparison))

        container = QWidget()
        container.setLayout(layout)
        self.setCentralWidget(container)

    def _button(self, text, callback):
        btn = QPushButton(text)
        btn.clicked.connect(callback)
        return btn

    def _select_file(self, label, attr):
        path, _ = QFileDialog.getOpenFileName(self, "Select BOM File", ".", "Excel Files (*.xlsx)")
        if path:
            setattr(self, attr, path)
            label.setText(f"{attr.replace('file', 'BOM File ')}: {os.path.basename(path)}")

    def select_file1(self):
        self._select_file(self.label_file1, 'file1')

    def select_file2(self):
        self._select_file(self.label_file2, 'file2')

    def run_comparison(self):
        if not self.file1 or not self.file2:
            QMessageBox.warning(self, "Error", "Select both BOM files first.")
            return
        try:
            os.makedirs(OUTPUT_DIR, exist_ok=True)
            output_file = self._output_name()
            df1, df2 = self._read_excel()
            df_merged = self._merge_and_filter(df1, df2)
            output_file = self._handle_existing(output_file)
            if not output_file:
                return
            self._write_excel(output_file, df1, df2, df_merged)
            self._format_excel(output_file)
            self._write_refdes_txt(df_merged)
            QMessageBox.information(self, "Success", f"Saved to {output_file}")
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))

    def _output_name(self):
        n1 = os.path.basename(self.file1).split('.')[0].split('_')[-1]
        n2 = os.path.basename(self.file2).split('.')[0].split('_')[-1]
        return os.path.join(OUTPUT_DIR, f'BOM_Differences_{n1}_vs_{n2}.xlsx')

    def _read_excel(self):
        df1 = pd.read_excel(self.file1, skiprows=15, skipfooter=7, usecols=COLUMNS).set_index('RefDes')
        df2 = pd.read_excel(self.file2, skiprows=15, skipfooter=7, usecols=COLUMNS).set_index('RefDes')
        return df1, df2

    def _merge_and_filter(self, df1, df2):
        df = pd.merge(df1, df2, how='outer', left_index=True, right_index=True, suffixes=('_BOM1', '_BOM2'))
        if self.cb_zero_qty.isChecked():
            df = df[~((df['Quantity_BOM1'] == 0) & (df['Quantity_BOM2'] == 0))]
        if self.cb_new_zero.isChecked():
            df = df[~((df['Quantity_BOM2'] == 0) & (df['Quantity_BOM1'].isna()))]
        if self.cb_dep_zero.isChecked():
            df = df[~((df['Quantity_BOM1'] == 0) & (df['Quantity_BOM2'].isna()))]
        if self.cb_unchanged.isChecked():
            df = df.loc[~(df.filter(like='_BOM1').fillna('').values ==
                          df.filter(like='_BOM2').fillna('').values).all(axis=1)]
        return df.sort_index()

    def _handle_existing(self, output_file):
        if not os.path.exists(output_file):
            return output_file
        box = QMessageBox(self)
        box.setWindowTitle("File Already Exists")
        box.setText(f"'{os.path.basename(output_file)}' already exists.")
        box.setInformativeText("What would you like to do?")
        btn_overwrite = box.addButton("Overwrite", QMessageBox.AcceptRole)
        btn_rename = box.addButton("Rename", QMessageBox.ActionRole)
        btn_cancel = box.addButton("Cancel", QMessageBox.RejectRole)
        box.exec_()
        clicked = box.clickedButton()
        if clicked == btn_overwrite:
            return output_file
        if clicked == btn_rename:
            ts = time.strftime("%d.%m.%y_%H%M")
            base = os.path.basename(output_file).split('.')[0]
            return os.path.join(OUTPUT_DIR, f'{base}_{ts}.xlsx')
        QMessageBox.information(self, "Cancelled", "Operation cancelled.")
        return None

    def _write_excel(self, output_file, df1, df2, df_merged):
        while True:
            try:
                with pd.ExcelWriter(output_file, mode='w') as writer:
                    pd.concat([df1, df2], axis=1, join='outer').sort_index().to_excel(
                        writer, sheet_name='Side by Side Comparison')
                    df_merged.to_excel(writer, sheet_name='BOM Delta')
                break
            except PermissionError:
                box = QMessageBox(self)
                box.setWindowTitle("Permission Error")
                box.setText("File is open in another process. Close it and retry.")
                btn_retry = box.addButton("Retry", QMessageBox.AcceptRole)
                btn_cancel = box.addButton("Cancel", QMessageBox.RejectRole)
                box.exec_()
                if box.clickedButton() == btn_cancel:
                    return
                time.sleep(1)

    def _format_excel(self, output_file):
        wb = load_workbook(output_file)
        ws = wb['Side by Side Comparison']
        ws.insert_rows(1)
        ws.merge_cells('B1:I1')
        ws.merge_cells('J1:Q1')
        ws['B1'], ws['J1'] = 'BOM1', 'BOM2'
        for cell in [ws['B1'], ws['J1']]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        ws.auto_filter.ref = 'A2:U2'
        wb.save(output_file)
        open_file(output_file)

    def _write_refdes_txt(self, df_merged):
        n1 = os.path.basename(self.file1).split('.')[0]
        n2 = os.path.basename(self.file2).split('.')[0]
        txt_path = os.path.join(OUTPUT_DIR, f'RefDesList_{n1}_vs_{n2}.txt')
        with open(txt_path, 'w') as f:
            f.writelines(f"{r}\n" for r in df_merged.index)


def main():
    app = QApplication(sys.argv)
    window = BOMComparisonApp()
    window.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
